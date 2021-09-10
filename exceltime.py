from openpyxl import load_workbook
import datetime
import holidays

# file stuff -> copy file, open copy and open with openpyxl
get_file_name = input("Name des Excelsheets eingeben (ohne .xlsx). Falls der Name 'Leistungsbericht' ist gib einfach y ein> ")
operation = input("Werkstudent oder Vollzeitkraft (w oder v)> ")

if str(get_file_name).lower() == "y":
    file_name = "Leistungsbericht"
else:
    file_name = str(get_file_name)

# file_name = "Leistungsbericht"
# operation = "w"

wb = load_workbook(filename=file_name + ".xlsx", data_only=True)

# declare variables we need
threshold_time = datetime.timedelta(hours=8, minutes=0, seconds=0)
german_holidays = holidays.Germany(prov="BW", years=[2020, 2021, 2022])


# if more than 1 entry for a day check how much and handle them different
def calculate_multiple_times(start_index, curr_sheet, start_date):
    skip_index = []
    for row in range(start_index, 10000):
        if curr_sheet.cell(row=row, column=1).value == start_date:
            skip_index.append(row)
        else:
            return skip_index, True


# calculate multiple entries for one day
def check_time_limit_multiple_days(curr_time, new_day):
    if new_day > threshold_time:
        return threshold_time - curr_time + datetime.timedelta(hours=0, minutes=30, seconds=0), False
    if curr_time + new_day >= threshold_time:
        possible_time = threshold_time - curr_time
        return possible_time, False

    else:
        return new_day, True


week_time = datetime.timedelta()
week_nr = 0


def calculate_one_month(active_sheet, curr_week, time_tank):

    delete_empty_rows = []

    # check for empty rows and delete them
    for row in range(5, 50):
        if active_sheet.cell(row=row, column=1).value is None and active_sheet.cell(row=row, column=8).value is None:
            delete_empty_rows.insert(0, row)

    for row_to_delete in delete_empty_rows:
        active_sheet.delete_rows(row_to_delete)

    # declare variables we need
    max_row = active_sheet.max_row
    time_multiple_days = datetime.timedelta(hours=0, minutes=0, seconds=0)
    real_time_month = datetime.timedelta(hours=0, minutes=0, seconds=0)
    can_still_work = True
    multiple_days_active = False
    rows_to_delete = []
    skip_rows = []

    # loop through rows in the excel sheet
    for row in range(4, max_row + 1):
        curr_cell_time = active_sheet.cell(row=row, column=6).value
        curr_cell_date = active_sheet.cell(row=row, column=1).value
        next_day = active_sheet.cell(row=row + 1, column=1).value

        if next_day is None:
            next_day = datetime.datetime(year=2000, month=1, day=1)

        if curr_cell_time is not None and curr_cell_date is not None:
            time_worked_day = curr_cell_time
            real_time_month += time_worked_day
            xy = active_sheet.cell(row=row, column=3).value
            leave_time = datetime.timedelta(hours=xy.hour, minutes=xy.minute, seconds=xy.second)

            # check if sunday or holiday
            if curr_cell_date.weekday() == 6 or curr_cell_date in german_holidays:
                active_sheet.cell(row=row, column=2).value = datetime.timedelta(hours=0, minutes=0, seconds=0)
                active_sheet.cell(row=row, column=3).value = datetime.timedelta(hours=0, minutes=0, seconds=0)
                continue

            if curr_cell_date == next_day and not multiple_days_active:
                skip_rows, multiple_days_active = calculate_multiple_times(row, active_sheet, curr_cell_date)

            # the calculation if multiple entries for one day
            if row in skip_rows:
                if not can_still_work:
                    rows_to_delete.insert(0, row)

                time_worked_day, can_still_work = check_time_limit_multiple_days(time_multiple_days, time_worked_day)
                time_multiple_days += time_worked_day

                if not can_still_work:
                    yx = active_sheet.cell(row=row, column=2).value
                    time_come = datetime.timedelta(hours=yx.hour, minutes=yx.minute, seconds=yx.second)
                    active_sheet.cell(row=row, column=3).value = time_come + time_worked_day

                if row + 1 not in skip_rows:
                    can_still_work = True
                    multiple_days_active = False
                    time_multiple_days = datetime.timedelta(hours=0, minutes=0, seconds=0)

                continue

            # check if limit is reached
            if curr_cell_time > threshold_time:

                additional_pause_time = datetime.timedelta(hours=0, minutes=0, seconds=0)

                # if pause 45 min remove 15, bec otherwise new calced time is too high
                if int(active_sheet.cell(row=row, column=4).value[3:]) > 30:
                    additional_pause_time = datetime.timedelta(hours=0, minutes=15, seconds=0)

                excessive_time = time_worked_day - threshold_time
                active_sheet.cell(row=row, column=3).value = leave_time - excessive_time - additional_pause_time

    # delete rows with multiple entries for one day if limit is reached
    for row in rows_to_delete:
        active_sheet.delete_rows(row)

    # add all constant fields in case they got removed
    active_sheet.cell(row=5, column=7).value = real_time_month
    active_sheet.cell(row=4, column=7).value = "=SUM(F4:F33)"
    active_sheet["G26"] = "Gesamt"
    active_sheet["G28"] = "Unterschrift Arbeitgeber:"

    # calc time for werkstudent (add 20 hours weeklimit)
    if operation == "w":

        for row in range(4, 60):
            curr_date = active_sheet.cell(row=row, column=1).value

            if curr_date is None:
                break

            if curr_date.isocalendar()[1] != curr_week:
                curr_week = curr_date.isocalendar()[1]
                time_tank = datetime.timedelta()

            # if new week starts reset time and update week

            time_tank = active_sheet.cell(row=row, column=6).value

            if time_tank == datetime.timedelta(hours=50):
                active_sheet.cell(row=row, column=6).value = datetime.timedelta()
                continue

            elif time_tank >= datetime.timedelta(hours=20):
                uy = active_sheet.cell(row=row, column=3).value
                go_time = datetime.timedelta(hours=uy.hour, minutes=uy.minute, seconds=uy.second)
                overlapping_time = time_tank - datetime.timedelta(hours=20)
                active_sheet.cell(row=row, column=3).value = go_time - overlapping_time
                time_tank = datetime.timedelta(hours=50)

    # add formulas again
    for row in range(4, 50):
        active_sheet.cell(row=row, column=6).value = f"=C{row}-B{row}-D{row}-E{row}"
        active_sheet.cell(row=row,
                          column=4).value = f'=IF(C{row}-B{row}>$H$4,"00:45",IF(C{row}-B{row}>$H$5,"00:30",IF(C{row}-B{row}<=$H$5,"00:00")))'

    return curr_week, time_tank


# loop through each sheet
for index, sheet in enumerate(wb.worksheets):
    week_nr, week_time = calculate_one_month(sheet, week_nr, week_time)
    print(f"{index + 1} von {len(wb.worksheets)}: Done")

wb.save(filename=file_name + "_fertig.xlsx")
print("_" * 30 + "\n")
print(f"Success\n{file_name}_fertig.xlsx ist fertig")
